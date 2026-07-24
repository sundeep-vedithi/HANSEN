#!/usr/bin/env python3
"""Build the archived HANSEN Database export data.

This program is derived from the HANSEN server extraction procedure under
`/data/new-hansen` on cam-server. It retains only the fields required for the
eight HANSEN TSV exports. Protein sequences, model coordinates, residue
predictions, logs and web application code are excluded.
"""

from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import sqlite3
from pathlib import Path
from typing import Any


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r", " ").replace("\n", " ").strip()
    return value


def worksheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Building the HANSEN data file requires openpyxl. Install it with python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(sheet.iter_rows(min_row=4, max_row=4))
    ]
    records: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=5, values_only=True):
        record = {
            header: clean(value)
            for header, value in zip(headers, values)
            if header
        }
        if record.get("ML ID"):
            records.append(record)
    workbook.close()
    return records


def metadata_records(database_path: Path) -> list[dict[str, Any]]:
    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            """
            SELECT
                ml_id,
                gene_names AS gene,
                protein_names AS protein,
                entry AS uniprot,
                pathway,
                gene_ontology_ids AS go_terms
            FROM protein_characteristics
            """
        ).fetchall()
    finally:
        connection.close()

    # The final database row for a duplicated ML identifier supplies the
    # display information. This agrees with the HANSEN server procedure.
    by_ml_id: dict[str, dict[str, Any]] = {}
    for row in rows:
        ml_id = str(row["ml_id"] or "").strip().upper()
        if not ml_id:
            continue
        by_ml_id[ml_id] = {
            "ML ID": ml_id,
            "Gene": clean(row["gene"]),
            "Protein": clean(row["protein"]),
            "UniProt": clean(row["uniprot"]),
            "Pathway": clean(row["pathway"]),
            "GO Terms": clean(row["go_terms"]),
        }
    return [by_ml_id[key] for key in sorted(by_ml_id)]


def hansen_metrics(
    database_path: Path,
    cache_metrics: list[dict[str, Any]],
    reference_workbook: Path,
) -> list[dict[str, Any]]:
    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            """
            SELECT
                ml_id,
                proteomelm_ess_probability,
                proteomelm_ess_class,
                proteomelm_ess_call
            FROM hansen_essentiality_predictions
            """
        ).fetchall()
    finally:
        connection.close()

    essentiality = {
        str(row["ml_id"]).strip().upper(): {
            "score": row["proteomelm_ess_probability"],
            "class": clean(row["proteomelm_ess_class"])
            or clean(row["proteomelm_ess_call"]).replace("-", " ").title(),
        }
        for row in rows
    }
    reference = {
        row["ML ID"]: row
        for row in worksheet_records(reference_workbook, "S8_Combined")
    }

    output = []
    for source in cache_metrics:
        row = dict(source)
        ml_id = str(row.get("ml_id") or "").strip().upper()
        if ml_id in essentiality:
            row["essentiality_score"] = essentiality[ml_id]["score"]
            row["essentiality_call"] = essentiality[ml_id]["class"]
        if ml_id in reference:
            # Preserve the validated model aggregation values when the live
            # statistics cache has since moved.
            row["mean_plddt"] = reference[ml_id]["Mean pLDDT"]
            row["assemblies"] = reference[ml_id]["Assemblies"]
        output.append(row)
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--cache", required=True, type=Path)
    parser.add_argument("--reference-workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    cache = json.loads(args.cache.read_text(encoding="utf-8"))
    metrics = cache.get("table_rows") or []
    if len(metrics) != 1603:
        raise SystemExit(
            f"Expected 1,603 protein measurement rows but found {len(metrics):,}"
        )

    model_records = worksheet_records(args.reference_workbook, "S1_All_Models")
    for record in model_records:
        record.pop("Example Selection Score", None)
    model_records.sort(key=lambda row: int(row["Database Row"]))

    ligand_records = worksheet_records(
        args.reference_workbook, "S3_Oligomer_Ligands"
    )

    snapshot = {
        "schema_version": "1.0",
        "dataset": "Archived HANSEN Database export data",
        "source": {
            "server_root": "/data/new-hansen",
            "database_file": "protein_characteristics.db",
            "database_sha256": sha256(args.database),
            "statistics_cache_file": (
                "static/statistics_cache/hansen_statistics_advanced_v4.json"
            ),
            "statistics_cache_sha256": sha256(args.cache),
            "statistics_cache_generated_at": cache.get("generated_at", ""),
            "reference_workbook_sha256": sha256(args.reference_workbook),
        },
        "scope": {
            "included": [
                "protein display metadata",
                "model records used in HANSEN exports S1 and S2",
                "curated oligomer ligand records used in HANSEN export S3",
                "protein measurements used in HANSEN exports S4 to S8",
            ],
            "excluded": [
                "amino acid sequences",
                "structure coordinate files",
                "residue pocket and epitope predictions",
                "server logs, credentials and web application code",
            ],
        },
        "metadata": metadata_records(args.database),
        "models": model_records,
        "oligomer_ligands": ligand_records,
        "protein_metrics": hansen_metrics(
            args.database, metrics, args.reference_workbook
        ),
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(args.output, "wt", encoding="utf-8", compresslevel=9) as handle:
        json.dump(snapshot, handle, ensure_ascii=False, separators=(",", ":"))
        handle.write("\n")

    print(f"Wrote {args.output}")
    print(f"Models {len(model_records):,}")
    print(f"Oligomer ligands {len(ligand_records):,}")
    print(f"Protein measurements {len(metrics):,}")
    print(f"Compressed size {args.output.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
